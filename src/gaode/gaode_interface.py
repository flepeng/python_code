# -*- coding:utf-8 -*-
"""
    @Author : Feng Lepeng
    @Date   : 2024/8/11 11:27
    @File   : gaode_interface.py
    @Desc   :
"""

import math
import re
import openpyxl
import requests
from bs4 import BeautifulSoup
from src.utils.logging_util import logger
import pymysql

logging = logger


def case_from_mysql(database, table):
    """从mysql中获取的case"""
    conn = pymysql.connect(host="127.0.0.1", port=3306, user="root", password="root", charset="utf8", db=database)
    cursor = conn.cursor()
    sql = "select * from " + table
    cursor.execute(sql)
    conn.commit()
    result = cursor.fetchall()

    # 获取数据表的title，cursor.description获取的是上一个cursor.execute的title
    desc = cursor.description
    title = [item[0] for item in desc]

    # 将title和case组成字典，方便调用
    cases = []
    for case in result:
        cases.append(dict(zip(title, case)))

    # 有的字典的值也是一个字典，但此时是字符串格式，因此要转换正真正的字典
    for case in cases:
        for key, value in case.items():
            # 此时的value有很多的引号，加一个str会去掉多余的引号，不然下面的正则表达式会报错，并且有的值是数据库的主键，即整数，也顺便转换成字符串
            value = str(value)
            if "{" in value and "}" in value and ":" in value:  # 判断值是否是字典，有冒号和大括号的就是字典形式的字符串
                value = re.findall(r"([\"\'])*([\w\d]+)(\1)|\"(\"{2})\"", value)  # 正则表达式去掉列字符串中的引号
                value = [i[1] for i in value]  # 取正则表达式中的第2个分组，也就是最终想要的结果
                # 列表的步长分别取出列表中的值当做字典中的key和value
                keys = value[::2]
                values = value[1::2]

                value = dict(zip(keys, values))  # 将key和value组成字典
                case[key] = value
    conn.close()
    cursor.close()
    return cases


# case_from_mysql("interface_cases", "cases")


def load_cases(filename, sheet):
    """从excel中获取测试用例"""
    wb = openpyxl.load_workbook(filename)
    ws = wb[sheet]
    cases = []
    # 获取excel中的数据，并将数据变成字典，方便后面调用，有两种方法
    # 方法一，此方法需要将用到的数据逐个添加到字典中，如果数据多，代码就会很长
    # for i in range(7, ws.max_row + 1):
    #     dict_case = dict(id=ws.cell(row=i, column=1).value,
    #                      header=ws.cell(row=i, column=3).value,
    #                      params=ws.cell(row=i, column=4).value,
    #                      status_code=ws.cell(row=i, column=6).value,
    #                      expected=ws.cell(row=i, column=7).value)
    #     cases.append(dict_case)
    # 方法二，此方法更灵活，更方便，一次性将excel中所有的数据都变成了字典
    title = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2):
        temp = []
        for cell in row:
            temp.append(cell.value)
        cases.append(dict(zip(title, temp)))
    return cases


def get_location(address, city=None):
    """获取经纬度坐标，供其他接口调用，例如路径规划接口"""
    url = "https://restapi.amap.com/v3/geocode/geo"
    params = {
        "key": "52991612454e98c850d4168f34bc46e2",
        "address": address,
        "city": city,
        # "sig": 'md5(address=南京东路&city=上海&key=52991612454e98c850d4168f34bc46e2&a921467fa2b7b6e4a43ee71921b54ab7)'
        # "sig": '5904e61b92dbcefacbc32d4bf61a304a'
    }
    res = requests.get(url=url, params=params)
    status_code = res.status_code
    result = res.json()
    # pprint(res.json())

    status = result["status"]
    info = result["info"]
    infocode = result["infocode"]
    if status == "1" and info == "OK" and infocode == "10000":
        geocodes = result.get("geocodes")[0]
        formatted_address = geocodes["formatted_address"]
        location = geocodes["location"]
        logging.info(f"匹配到的地址为：{formatted_address}")
        logging.info(f"经纬度为：{location}")
        return location
    else:
        logging.info(f"没有匹配到'{city}'的'{address}'，请检查后重新输入！")
        logging.info(f"状态码：{status_code}, {status}, {info}, {infocode}")


def request(case_name: str, case: dict, url: str, params: dict) -> tuple:
    """接口请求封装"""
    expect_result = case.get("expected")
    logging.info(f"用例{case_name}")
    logging.info(f"{case_name}--请求参数：{params}")
    logging.info(f"{case_name}--预期的状态码：{expect_result}")

    logging.info(f"{case_name}--开始发送接口请求")
    res = requests.get(url=url, params=params)
    status_code = res.status_code
    logging.info(f"{case_name}--通信状态码：{status_code}")
    logging.info(f"{case_name}--开始获取响应结果")

    response = res.text
    # output参数是由用户输入的，它可指定返回的数据格式，有json和xml两种，所以此处加一个判断，
    # 如果是xml格式，字符串中会有xml开头的标题
    # print("判断返回的数据格式是否是xml？", bool("xml" in result))
    if "xml" in response:
        # 然后用BeautifulSoup模块解析出数据中的status、inof、infocode等的值
        logging.info(f"{case_name}--响应数据格式：XML")
        soup = BeautifulSoup(response, features="xml")
        logging.info(f"{case_name}--提取响应头")
        status = soup.select_one("status").text
        info = soup.select_one("info").text
        infocode = soup.select_one("infocode").text
        count = soup.select_one("count").text
        logging.info(f"{case_name}--响应状态码：{info}, {infocode}, {status}, {count}")
    else:
        # 如果是json字符串格式，就用json()方法转化为字典，然后取status、info、infocode等的值
        logging.info(f"{case_name}--响应数据格式：Json")
        response = res.json()
        logging.info(f"{case_name}--提取响应头")
        status = response["status"]
        info = response["info"]
        infocode = response["infocode"]
        count = response.get("count")  # 不知道为什么这里不能用response["count"]，报keyError
        logging.info(f"{case_name}--响应状态码：{status}, {info}, {infocode}, {count}")
    return status_code, status, info, infocode, response


class GaodeInterface:
    def geo(self, interface: str, case: dict) -> tuple:
        """地理编码接口"""
        # 获取case的一些基本信息
        # 从excel中获取的params是字符串类型，用eval变成字典
        url = "https://restapi.amap.com" + interface
        case_name = str(case.get("id")) + "." + case.get("case_name")
        # 如果是从本地的excel中读取的case就是字符串形式的字典，用eval变成真正的字典
        if isinstance(case.get("params"), str):
            params = eval(case.get("params"))
        else:
            # 如果是从mysql中获取的case，本身就是真正的字典，不需要转换
            params = case.get("params")
        status_code = request(case_name=case_name, case=case, url=url, params=params)
        return status_code

    def walking_path(self, interface: str, case: dict, city=None) -> tuple:
        """路径规划接口
            输入出发地和目的地会计算出它们中间的路线
            :param interface: 测试接口
            :param case: 测试用例
            :param city: 默认在全国范围查找，输入city会使匹配的地址会更准确
        """
        url = "https://restapi.amap.com" + interface
        case_name = str(case.get("id")) + "." + case.get("case_name")
        # 如果是从本地的excel中读取的case就是字符串形式的字典，用eval变成真正的字典
        if isinstance(case.get("params"), str):
            params = eval(case.get("params"))
        else:
            # 如果是从mysql中获取的case，本身就是真正的字典，不需要转换
            params = case.get("params")

        # 此接口的出发地和目的地只能是经纬度形式，所以先把excel中的中文地址提取出来 再用函数转化成经纬度形式
        origin = params.get("origin")
        destination = params.get("destination")
        params["origin"] = get_location(origin, city)  # 转化成经纬度
        params["destination"] = get_location(destination, city)

        # 发送请求，并得到响应码和响应内容
        status_code = request(case_name=case_name, case=case, url=url, params=params)

        # 提取响应内容
        response = status_code[4]
        status, info, infocode = status_code[1], status_code[2], status_code[3]
        if status == "1" and info == "ok" and infocode == "10000":
            # 处理xml格式的数据
            if "xml" in response:
                soup = BeautifulSoup(response, features="xml")
                distance = soup.select_one("path>distance").text
                duration = math.ceil(int(soup.select_one("path>duration").text) / 60)
                steps_list = soup.select("steps>step")
                logging.info(f"{origin} --> {destination}")
                logging.info(f"距离：{distance}米")
                logging.info(f"步行时间约：{duration}分钟")
                for step in steps_list:
                    step_duration = step.select_one("duration").text
                    step_instruction = step.select_one("instruction").text
                    logging.info(f"-> {step_instruction}, 大约需要{math.ceil(int(step_duration) / 60)}分钟")
                return status_code
            # 处理json格式的数据
            else:
                paths = response.get("route").get("paths")
                distance = paths[0].get("distance")
                duration = math.ceil(int(paths[0].get("duration")) / 60)
                steps_list = paths[0].get("steps")
                logging.info(f"{origin} --> {destination}")
                logging.info(f"距离：{distance}米")
                logging.info(f"步行时间约：{duration}分钟")
                for step in steps_list:
                    step_duration = step.get("duration")
                    step_instruction = step.get("instruction")
                    logging.info(f"-> {step_instruction}, 大约需要{math.ceil(int(step_duration) / 60)}分钟, ")
                return status_code
        else:
            return status_code

    def weather(self, interface, case):
        url = "https://restapi.amap.com" + interface
        case_name = str(case.get("id")) + "." + case.get("case_name")
        # 如果是从本地的excel中读取的case就是字符串形式的字典，用eval变成真正的字典
        if isinstance(case.get("params"), str):
            params = eval(case.get("params"))
        else:
            # 如果是从mysql中获取的case，本身就是真正的字典，不需要转换
            params = case.get("params")
        status_code = request(case_name=case_name, case=case, url=url, params=params)
        response = status_code[4]
        status, info, infocode = status_code[1], status_code[2], status_code[3]

        # 由于无论city无论输入的什么，status、info、infocode他们都不会报错，
        # 因此多添加一个count判断，如果count=0什么输入有误，并且加个if来分别获取json和xml格式所返回的count
        if "xml" in response:
            count = re.search("<count>(\d)</", response).group(1)
        else:
            count = response.get("count")
        if status == "1" and info == "OK" and infocode == "10000" and int(count) > 0:
            if "xml" in response:
                soup = BeautifulSoup(response, features="xml")
                province = soup.find("province").text
                city = soup.find("city").text
                reporttime = soup.find("reporttime").text
                adcode = soup.find("adcode").text
                weather = soup.find("weather").text
                temperature = soup.find("temperature").text
                winddirection = soup.find("winddirection").text
                windpower = soup.find("windpower").text
                humidity = soup.find("humidity").text
                logging.info(f"天气预报：\n{province}, {city}, {adcode}\n"
                             f"{reporttime} \n天气：{weather} \n温度：{temperature} \n风向：{winddirection} \n风力：{windpower} \n湿度：{humidity}")
                return status_code
            else:
                lives = response.get("lives")[0]
                province = lives.get("province")
                city = lives.get("city")
                adcode = lives.get("adcode")
                reporttime = lives.get("reporttime")
                weather = lives.get("weather")
                temperature = lives.get("temperature")
                winddirection = lives.get("winddirection")
                windpower = lives.get("windpower")
                humidity = lives.get("humidity")
                logging.info(f"天气预报：\n{province}, {city}, {adcode}\n"
                             f"{reporttime} \n天气：{weather} \n温度：{temperature} \n风向：{winddirection} \n风力：{windpower} \n湿度：{humidity}")
                return status_code
        else:
            return status_code


if __name__ == '__main__':

    # load_cases("./cases.xlsx", "Geo")
    print(get_location("huishancun", "jiujiang"))
    gaode = GaodeInterface()
    interface = "/v3/weather/weatherInfo"
    case = {"id": "1", "case": "正例_正确的key",
            "params": '{"key": "52991612454e98c850d4168f34bc46e2", "city": "360429", "extensions": "base", "output": "json"}'}
    a = gaode.weather(interface, case)
    print(a)